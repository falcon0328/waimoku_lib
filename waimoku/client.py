import openpyxl as excel
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from . import WaimokuUser


class WaimokuClient:
    __merge_target_cell = "B1"
    __merge_target_cells = "B1:D1"
    __headerTitles: dict = {"A2": "通し番号", "B2": "区分", "C2": "氏名", "D2": "所属"}
    __header_collor: PatternFill = PatternFill(patternType='solid', fgColor="FFF2CC")
    __font_sizes: list = [9, 9, 11, 11]
    __cellWidths: list = [6.83, 4.83, 17.33, 44.67]
    __cell_side = Side(style='thin', color='000000')
    __cell_border: Border

    def __init__(self, *args, **kwargs):
        Border()
        self.__cell_border = Border(top=self.__cell_side, bottom=self.__cell_side, left=self.__cell_side, right=self.__cell_side)
        return super().__init__(*args, **kwargs)

    def save_to_file(self, user_list: [WaimokuUser], save_filename="event_participantsList.xlsx"):
        # ブックの新規作成
        wb = excel.Workbook()
        # アクティブなシートの作成
        ws = wb.active
        # シートタイトルはSheet1に変更
        ws.title = "Sheet1"
        # ヘッダーを用意
        for key in self.__headerTitles.keys():
            self.write(ws=ws, key=key, value=self.__headerTitles[key], fill=self.__header_collor, font_size=9)
        # シートにデータを書き込む
        for index, user in enumerate(user_list):
            self.write(ws=ws, key="A{0}".format(index + 3), value=index+1, font_size=9)
            self.write(ws=ws, key="C{0}".format(index + 3), value=user.full_name, font_size=11)
            self.write(ws=ws, key="D{0}".format(index + 3), value=user.assign, font_size=11)
        # シートの幅を調節する
        self.__adjust_sheet_size(ws)
        # 全てのセルに上下中央揃えを実施
        self.__adjust_sheet_alignment(ws)
        # 「入力必須」の表示を行うセルを用意
        # B1 ~ D1のセルを連結
        self.__merge_cells_and_setting_to_align(ws=ws, key=self.__merge_target_cell, merge_cells=self.__merge_target_cells, value="入力必須", fill=self.__header_collor, border=self.__cell_border)
        # ファイル保存
        wb.save(save_filename)

    def write(self, ws, key: str, value, fill: PatternFill = PatternFill(), border=Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000')), font_size: int = 9):
        """ワークシートの指定したキーにデータを書き込む

        Arguments:
            ws {} -- 対象のワークシート
            key {str} -- キー
            value {} -- 書き込む内容

        Keyword Arguments:
            fill {PatternFill} -- 色の指定 (default: {PatternFill()})
            border {Border} -- 罫線情報(default: {Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))})
            font_size {int} -- フォントサイズ (default: {9})
        """
        ws[key] = value
        ws[key].fill = fill
        ws[key].border = border
        ws[key].font = Font(size=font_size)

    def __adjust_sheet_size(self, ws):
        """指定したシートの幅を調節する

        Arguments:
            ws {[]} -- 対象のシート
        """
        for index, col in enumerate(self.__cellWidths):
            ws.column_dimensions[excel.utils.get_column_letter(index + 1)].width = self.__cellWidths[index]
        for row in range(ws.max_row + 1):
            ws.row_dimensions[row].height = 27

    def __adjust_sheet_alignment(self, ws, align: Alignment = Alignment(vertical="center")):
        """セルの値を上下左右、指定したアラインにする

        Arguments:
            ws {[type]} -- 対象のシート

        Keyword Arguments:
            align {Alignment} -- アライン情報 (default: {Alignment(vertical="center")})
        """
        for row in ws:
            for cell in row:
                cell.alignment = align

    def __merge_cells_and_setting_to_align(self, ws, key: str, merge_cells: str, value, fill: PatternFill = PatternFill(), align: Alignment = Alignment(horizontal="center", vertical="center"), border=Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))):
        """指定したセル達を連結後に値をセットし、アラインと色を適用する

        Arguments:
            ws {} -- 対象のワークシート
            key {str} -- 対象のセル番号
            merge_cells {str} -- 連結させたいセル
            value {[type]} -- 書き込む内容

        Keyword Arguments:
            fill {PatternFill} -- 色の指定 (default: {PatternFill()})
            align {Alignment} -- アライン情報 (default: {Alignment(horizontal="center", vertical="center")})
            border {Border} -- 罫線情報(default: {Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))})
        """
        ws[key] = value
        ws[key].border = border
        ws.merge_cells(merge_cells)
        ws[key].fill = fill
        ws[key].alignment = align
